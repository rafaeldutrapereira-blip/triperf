"""
Lab report parser — extracts biomarker values from PDF/CSV/XLSX text.
Handles common Chilean/Spanish laboratory report formats.
"""
from __future__ import annotations

import io
import re
import unicodedata
from typing import Any


# ---------------------------------------------------------------------------
# Synonym dictionary — each key maps to a list of Spanish/English text patterns
# All patterns are accent-free (text is normalized before matching)
# ---------------------------------------------------------------------------
SYNONYMS: dict[str, list[str]] = {
    "glucosa":           ["glucosa", "glucose"],
    "nitrogeno_ureico":  ["nitrogeno ureico", "nitrogen ureico", "bun",
                          "nitrogeno urei"],
    "urea":              ["urea"],
    "colesterol_total":  ["colesterol total", "colesterol (total)", "cholesterol total",
                          "col total"],
    "ac_urico":          ["acido urico", "urico", "urate", "acido urico (sangre)"],
    "proteinas_totales": ["proteinas totales", "proteina total", "proteins totales"],
    "albumina":          ["albumina", "albumin"],
    "globulinas":        ["globulinas", "globulin"],
    "bilirrubina_total": ["bilirrubina total", "bilirubin total", "bili total"],
    "got_ast":           ["got", "ast", "aspartato aminotransferasa",
                          "transaminasa oxalacetica", "transaminasa oxalactica",
                          "got (ast)", "ast (got)"],
    "gpt_alt":           ["gpt", "alt", "alanino aminotransferasa",
                          "transaminasa piruvica", "gpt (alt)", "alt (gpt)"],
    "ggt":               ["ggt", "gamma glutamil transferasa",
                          "gamma-glutamil", "gamma glutamil"],
    "ldh":               ["ldh", "deshidrogenasa lactica",
                          "lactato deshidrogenasa", "lactate dehydrogenase"],
    "fosfatasas_alc":    ["fosfatasa alcalina", "fosfatasas alcalinas",
                          "alkaline phosphatase", "alp"],
    "calcio":            ["calcio", "calcium", "calcio total"],
    "fosforo":           ["fosforo", "phosphorus", "fosforo inorganico"],
    "trigliceridos":     ["trigliceridos", "triglycerides", "trigliceridos (vldl)"],
    "hdl":               ["hdl", "colesterol hdl", "hdl colesterol", "hdl-c",
                          "hdl (colesterol bueno)"],
    "ldl":               ["ldl", "colesterol ldl", "ldl colesterol", "ldl-c",
                          "ldl (colesterol malo)"],
    "ck":                ["creatinquinasa", "creatinkinasa", "creatina quinasa",
                          "creatinkinase", "ck total", "cpk", "cpk total",
                          "creatinina quinasa"],
    "vitamina_b12":      ["vitamina b12", "b12", "cobalamina",
                          "cianocobalamina", "vitamina b 12"],
    "creatinina":        ["creatinina", "creatinine", "creatinina serica"],
    "tfge":              ["tfge", "filtrado glomerular", "tasa de filtracion",
                          "egfr", "tasa filtracion glomerular"],
    "psa_total":         ["psa total", "psa", "antigeno prostatico especifico"],
    "tsh":               ["tsh", "tirotropina", "hormona estimulante tiroides"],
    "ft4":               ["ft4", "t4 libre", "tiroxina libre", "t4l", "free t4"],
    "na":                ["sodio", "sodium", "na (sodio)", "sodio (na)"],
    "k":                 ["potasio", "potassium", "k (potasio)", "potasio (k)"],
    "cl":                ["cloro", "cloruro", "chloride", "cloro (cl)", "cl (cloro)"],
    "vitamina_d":        ["vitamina d total", "vitamina d", "25-oh vitamina d",
                          "25-hidroxivitamina d", "calcidiol", "25 oh vitamina d",
                          "vit d total", "vit. d"],
    "eritrocitos":       ["eritrocitos", "globulos rojos", "hematies",
                          "recuento eritrocitos", "recuento de eritrocitos",
                          "red blood cells", "rbc"],
    "hemoglobina":       ["hemoglobina", "hemoglobin", "hb", "hgb"],
    "hematocrito":       ["hematocrito", "hematocrit", "hto"],
    "vcm":               ["vcm", "volumen corpuscular medio", "mcv",
                          "vol corp medio"],
    "hcm":               ["hcm", "hemoglobina corpuscular media", "mch",
                          "hemoglobina corp media"],
    "chcm":              ["chcm", "concentracion hemoglobina", "mchc",
                          "conc hb corpuscular"],
    "leucocitos":        ["leucocitos", "globulos blancos", "white blood cells",
                          "wbc", "recuento leucocitos", "recuento de leucocitos"],
    "plaquetas":         ["plaquetas", "platelets", "trombocitos",
                          "recuento plaquetas", "recuento de plaquetas"],
    "sedimentacion":     ["sedimentacion", "vsg", "velocidad sedimentacion",
                          "erythrocyte sedimentation", "ves"],
}

# Reverse-lookup: pattern → key (built once at module load)
_PATTERN_TO_KEY: dict[str, str] = {
    p: k for k, patterns in SYNONYMS.items() for p in patterns
}


# ---------------------------------------------------------------------------
# Text normalization
# ---------------------------------------------------------------------------
def _normalize(text: str) -> str:
    """Lowercase + strip accents + normalize whitespace."""
    nfkd = unicodedata.normalize("NFKD", text.lower())
    no_acc = "".join(c for c in nfkd if not unicodedata.combining(c))
    # Normalize decimal separators: 17,0 → 17.0
    no_acc = re.sub(r"(\d),(\d)", r"\1.\2", no_acc)
    return no_acc


# ---------------------------------------------------------------------------
# Extraction helpers
# ---------------------------------------------------------------------------
_NUMBER_RE = r"(\d+\.?\d*)"


def _extract_value_from_text(normalized_text: str, synonyms: list[str]) -> float | None:
    """
    Search `normalized_text` for any synonym, then capture the first number
    within 70 characters after it (same or next line).
    Falls back to looking for number BEFORE the synonym (some table formats).
    """
    for syn in synonyms:
        # Forward: synonym → value
        fwd = rf'\b{re.escape(syn)}\b[^0-9\n]{{0,70}}{_NUMBER_RE}'
        m = re.search(fwd, normalized_text)
        if m:
            return float(m.group(1))
        # Also try line-by-line: "synonym   86   mg/dL"
        pattern_line = rf'^.*\b{re.escape(syn)}\b.*?{_NUMBER_RE}'
        m = re.search(pattern_line, normalized_text, re.MULTILINE)
        if m:
            return float(m.group(1))
    return None


# ---------------------------------------------------------------------------
# Public: PDF → dict
# ---------------------------------------------------------------------------
def extract_text_pdf(pdf_bytes: bytes) -> str:
    """
    Extract all text from a PDF using pdfplumber.
    Prioritizes table-based extraction for lab report grids.
    """
    import pdfplumber

    parts: list[str] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            # Table extraction (structured lab reports)
            tables = page.extract_tables() or []
            for table in tables:
                for row in table:
                    if row:
                        clean_row = [str(c).strip() if c else "" for c in row]
                        parts.append("  ".join(clean_row))
            # Raw text for any non-table content
            raw = page.extract_text(layout=False) or ""
            parts.append(raw)
    return "\n".join(parts)


def parse_lab_values(text: str) -> dict[str, float]:
    """
    Parse lab report text and return {biomarker_key: float_value}.
    Only returns keys that were successfully matched.
    """
    norm = _normalize(text)
    results: dict[str, float] = {}
    for key, synonyms in SYNONYMS.items():
        val = _extract_value_from_text(norm, synonyms)
        if val is not None:
            results[key] = val
    return results


# ---------------------------------------------------------------------------
# Public: CSV → dict
# ---------------------------------------------------------------------------
def parse_csv_bytes(csv_bytes: bytes) -> dict[str, float]:
    """
    Parse a CSV file.
    Expected columns: 'marcador' (or 'key') + 'valor' (or 'value').
    Also accepts the downloadable template format.
    """
    import pandas as pd

    try:
        df = pd.read_csv(io.BytesIO(csv_bytes), dtype=str)
    except Exception:
        return {}

    df.columns = [c.strip().lower() for c in df.columns]
    key_col = next((c for c in df.columns if c in ("marcador", "key", "marker")), None)
    val_col = next((c for c in df.columns if c in ("valor", "value", "resultado")), None)
    if not key_col or not val_col:
        return {}

    results: dict[str, float] = {}
    for _, row in df.iterrows():
        k = str(row[key_col]).strip().lower()
        v = str(row[val_col]).strip().replace(",", ".")
        if k in SYNONYMS and v and v not in ("nan", "", "-", "—"):
            try:
                results[k] = float(v)
            except ValueError:
                pass
    return results


# ---------------------------------------------------------------------------
# Public: XLSX → dict
# ---------------------------------------------------------------------------
def parse_xlsx_bytes(xlsx_bytes: bytes) -> dict[str, float]:
    """Parse an Excel file — same column convention as CSV."""
    import pandas as pd

    try:
        df = pd.read_excel(io.BytesIO(xlsx_bytes), dtype=str)
    except Exception:
        return {}

    df.columns = [c.strip().lower() for c in df.columns]
    key_col = next((c for c in df.columns if c in ("marcador", "key", "marker")), None)
    val_col = next((c for c in df.columns if c in ("valor", "value", "resultado")), None)
    if not key_col or not val_col:
        return {}

    results: dict[str, float] = {}
    for _, row in df.iterrows():
        k = str(row[key_col]).strip().lower()
        v = str(row[val_col]).strip().replace(",", ".")
        if k in SYNONYMS and v and v not in ("nan", "", "-", "—"):
            try:
                results[k] = float(v)
            except ValueError:
                pass
    return results


# ---------------------------------------------------------------------------
# Public: generate downloadable template
# ---------------------------------------------------------------------------
def generate_csv_template(biomarkers: dict) -> bytes:
    """Return a CSV template bytes with all biomarker keys pre-filled as rows."""
    import pandas as pd

    rows = [
        {
            "marcador":   key,
            "etiqueta":   bm["label"],
            "valor":      "",
            "unidad":     bm["unit"],
            "referencia": f"{bm['ref_low']}–{bm['ref_high']}",
        }
        for key, bm in biomarkers.items()
    ]
    return pd.DataFrame(rows).to_csv(index=False).encode("utf-8")


def generate_xlsx_template(biomarkers: dict) -> bytes:
    """Return a formatted Excel template bytes."""
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    rows = [
        {
            "marcador":   key,
            "etiqueta":   bm["label"],
            "valor":      None,
            "unidad":     bm["unit"],
            "referencia": f"{bm['ref_low']}–{bm['ref_high']}",
            "nota":       bm.get("athlete_note", ""),
        }
        for key, bm in biomarkers.items()
    ]
    df = pd.DataFrame(rows)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Examen")
        ws = writer.sheets["Examen"]

        # Header style
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center")

        # Highlight "valor" column (C) for user entry
        val_fill = PatternFill("solid", fgColor="FFF9E6")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.fill = val_fill
                cell.border = thin

        # Column widths
        for col, width in zip("ABCDEF", [22, 28, 12, 10, 16, 55]):
            ws.column_dimensions[col].width = width

    return buf.getvalue()
